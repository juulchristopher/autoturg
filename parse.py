#!/usr/bin/env python3
"""
Fetch the latest Transpordiamet infoleht xlsx and parse all 3 market categories
(järelturg, new cars, imports) into data.json.
Merges new month into existing data.json so history accumulates.
"""

import json, re, sys, os
from datetime import date, timedelta
from pathlib import Path
from urllib.request import urlopen, Request
from urllib.error import URLError, HTTPError

import openpyxl

MONTHS_EN = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']
BASE_URL = "https://www.transpordiamet.ee/sites/default/files/documents"
DATA_FILE = Path(__file__).resolve().parent / "data.json"

SKIP_MAKES = {'KOKKU', 'TOTAL', 'ZUSAMMEN', 'SUM', 'MARK', 'MÄRK'}

# Sheet detection keywords per category
SHEET_KEYWORDS = {
    'jarelturg': ('järelturg', 'jarelturg', 'omaniku', 'owner', 'used', 'kasutatud'),
    'newCars':   ('esmane', 'esmased', 'uued', 'uus', 'new', 'first registration', 'esmas'),
    'imports':   ('import', 'sisseveo', 'sissetoo'),
}


def candidate_urls(data_month: int, data_year: int) -> list:
    """Generate candidate URLs for a given data month.
    The file for data month M is uploaded in folder month M+1."""
    mm = str(data_month).zfill(2)
    # Upload folder is one month after the data month
    upload_month = data_month + 1
    upload_year = data_year
    if upload_month > 12:
        upload_month = 1
        upload_year += 1
    um = str(upload_month).zfill(2)

    urls = []
    # All known filename patterns
    suffixes = ['', '_statistika_esmased_ja_uued', '_stat_esm_ja_uued']
    # All known folder patterns: upload month, or data month
    folders = [f"{upload_year}-{um}", f"{data_year}-{mm}"]

    for folder in folders:
        for suffix in suffixes:
            for ext in ('.xlsx', '.xls'):
                urls.append(f"{BASE_URL}/{folder}/INFOLEHT-{mm}{data_year}{suffix}{ext}")
        # URL-encoded pattern: "Infoleht MM-YYYY (esmaste ja uute sõidukite statistika).xls"
        encoded_name = f"Infoleht%20{mm}-{data_year}%20%28esmaste%20ja%20uute%20s%C3%B5idukite%20statistika%29.xls"
        urls.append(f"{BASE_URL}/{folder}/{encoded_name}")
    # Also try with _0 suffix (seen for Jun 2024)
    for folder in folders:
        for ext in ('.xlsx', '.xls'):
            urls.append(f"{BASE_URL}/{folder}/INFOLEHT-{mm}{data_year}_statistika_esmased_ja_uued_0{ext}")

    return urls


def download(url: str) -> bytes:
    req = Request(url, headers={"User-Agent": "Mozilla/5.0 (autoturg-bot)"})
    with urlopen(req, timeout=60) as resp:
        return resp.read()


def find_sheet_by_category(wb, category: str):
    """Return the worksheet matching the given category keywords.
    Returns None if not found."""
    keywords = SHEET_KEYWORDS[category]
    priority = []
    rest = []
    for name in wb.sheetnames:
        lo = name.lower()
        if any(k in lo for k in keywords):
            priority.append(name)
        else:
            rest.append(name)

    for name in priority:
        ws = wb[name]
        if ws.max_row and ws.max_row >= 3:
            # Verify it has a header with make column
            for row in ws.iter_rows(min_row=1, max_row=min(25, ws.max_row), values_only=False):
                for cell in row:
                    val = str(cell.value or '').lower().strip()
                    if val in ('mark', 'märk', 'make'):
                        return ws
    return None


def find_jarelturg_sheet(wb):
    """Legacy wrapper — find järelturg sheet."""
    return find_sheet_by_category(wb, 'jarelturg')


# Model names where the first word alone is ambiguous and needs the second word
MULTI_WORD_MODELS = {'MODEL'}  # e.g. Tesla "MODEL S", "MODEL 3", "MODEL Y"


def split_model_variant(full_model: str) -> tuple:
    """Split a full model string into (model, variant).
    Handles multi-word model names like Tesla MODEL S, MODEL 3, etc."""
    parts = full_model.split()
    if not parts:
        return ('', '')
    if len(parts) >= 2 and parts[0] in MULTI_WORD_MODELS:
        model = f"{parts[0]} {parts[1]}"
        variant = ' '.join(parts[2:])
    else:
        model = parts[0]
        variant = ' '.join(parts[1:])
    return (model, variant)


def parse_sheet(ws) -> list:
    """Parse a worksheet into rows of {make, model, variant, fullModel, prodYear, count}.
    Works for all 3 categories since column structure is the same."""
    rows_out = []
    header_row = None
    make_col = model_col = prod_col = count_col = None

    for ri, row in enumerate(ws.iter_rows(min_row=1, max_row=min(25, ws.max_row), values_only=True), start=1):
        cells = [str(c or '').lower().strip() for c in row]
        mk_idx = next((i for i, c in enumerate(cells) if c in ('mark', 'märk', 'make')), None)
        if mk_idx is not None:
            header_row = ri
            make_col = mk_idx
            model_col = next((i for i, c in enumerate(cells) if i != mk_idx and (c in ('mudel', 'model') or 'mudel' in c)), None)
            prod_col = next((i for i, c in enumerate(cells) if 'aasta' in c or c == 'year' or 'tootmis' in c or 'esm reg' in c), None)
            count_col = next((i for i, c in enumerate(cells) if c in ('arv', 'kokku', 'hulk', 'transactions') or 'count' in c), None)
            if count_col is None:
                count_col = len(cells) - 1
            break

    if header_row is None or make_col is None:
        return []

    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row, values_only=True):
        vals = list(row)
        make = str(vals[make_col] or '').strip().upper()
        if not make or make in SKIP_MAKES:
            continue

        full_model = ''
        if model_col is not None and model_col < len(vals):
            full_model = str(vals[model_col] or '').strip().upper()

        model, variant = split_model_variant(full_model)

        prod_year = None
        if prod_col is not None and prod_col < len(vals):
            try:
                py = int(float(vals[prod_col]))
                if 1950 <= py <= date.today().year + 1:
                    prod_year = py
            except (ValueError, TypeError):
                pass

        count = 0
        if count_col is not None and count_col < len(vals):
            try:
                count = int(float(str(vals[count_col] or '0').replace(' ', '').replace(',', '')))
            except (ValueError, TypeError):
                pass

        if count > 0:
            rows_out.append({
                "make": make,
                "model": model,
                "variant": variant,
                "fullModel": full_model,
                "prodYear": prod_year,
                "count": count,
            })

    return rows_out


def load_data() -> dict:
    if DATA_FILE.exists():
        with open(DATA_FILE) as f:
            data = json.load(f)
        # Migrate old format: { months: [] } → { jarelturg: [], newCars: [], imports: [] }
        if "months" in data and "jarelturg" not in data:
            print("  Migrating old data format to multi-category...")
            return {
                "jarelturg": data["months"],
                "newCars": [],
                "imports": [],
            }
        # Ensure all keys exist
        data.setdefault("jarelturg", [])
        data.setdefault("newCars", [])
        data.setdefault("imports", [])
        return data
    return {"jarelturg": [], "newCars": [], "imports": []}


def save_data(data: dict):
    for key in ('jarelturg', 'newCars', 'imports'):
        data[key].sort(key=lambda m: m["year"] * 100 + m["month"])
    with open(DATA_FILE, "w") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def open_workbook(tmp_path, is_xls: bool):
    """Open an xlsx or xls file and return an openpyxl Workbook."""
    if is_xls:
        try:
            import xlrd
            xls_wb = xlrd.open_workbook(str(tmp_path))
            from openpyxl import Workbook
            wb = Workbook()
            for si, sheet in enumerate(xls_wb.sheets()):
                ws = wb.active if si == 0 else wb.create_sheet()
                ws.title = sheet.name
                for r in range(sheet.nrows):
                    for c in range(sheet.ncols):
                        ws.cell(row=r+1, column=c+1, value=sheet.cell_value(r, c))
            return wb
        except ImportError:
            print("  xlrd not installed, cannot read .xls files")
            return None
    else:
        return openpyxl.load_workbook(tmp_path, read_only=True, data_only=True)


def fetch_month(month: int, year: int) -> bool:
    """Try to fetch and parse a single month (all 3 categories). Returns True if any category succeeded."""
    print(f"Fetching infoleht for {MONTHS_EN[month-1]} {year}...")
    urls = candidate_urls(month, year)

    raw = None
    used_url = None
    for url in urls:
        try:
            raw = download(url)
            if len(raw) < 1000:
                raw = None
                continue
            used_url = url
            print(f"  Downloaded: {url}")
            break
        except (URLError, HTTPError):
            continue

    if raw is None:
        print(f"  No file found for {MONTHS_EN[month-1]} {year}")
        return False

    is_xls = used_url.endswith('.xls')
    tmp = DATA_FILE.parent / f"_tmp_infoleht{'.xls' if is_xls else '.xlsx'}"
    tmp.write_bytes(raw)

    try:
        wb = open_workbook(tmp, is_xls)
        if wb is None:
            return False

        print(f"  Sheets found: {wb.sheetnames}")
        data = load_data()
        any_success = False

        # Parse all 3 categories from the same workbook
        for category, data_key in [('jarelturg', 'jarelturg'), ('newCars', 'newCars'), ('imports', 'imports')]:
            ws = find_sheet_by_category(wb, category)
            if ws is None:
                print(f"  No {category} sheet found")
                continue

            rows = parse_sheet(ws)
            if not rows:
                print(f"  No data rows for {category} in sheet '{ws.title}'")
                continue

            print(f"  {category}: parsed {len(rows)} rows from sheet '{ws.title}'")

            # Upsert: remove existing month, add new
            data[data_key] = [m for m in data[data_key] if not (m["year"] == year and m["month"] == month)]
            data[data_key].append({
                "year": year,
                "month": month,
                "label": f"{MONTHS_EN[month-1]} {year}",
                "sheetUsed": ws.title,
                "rows": rows,
            })
            any_success = True

        if any_success:
            save_data(data)
            counts = {k: len(data[k]) for k in ('jarelturg', 'newCars', 'imports')}
            print(f"  Saved to {DATA_FILE} (jarelturg: {counts['jarelturg']}, newCars: {counts['newCars']}, imports: {counts['imports']} months)")

        if hasattr(wb, 'close'):
            wb.close()

        return any_success

    finally:
        tmp.unlink(missing_ok=True)


def main():
    today = date.today()
    # Fetch the previous month's data (published ~20th of current month)
    target = today.replace(day=1) - timedelta(days=1)
    month, year = target.month, target.year

    if not fetch_month(month, year):
        sys.exit(1)


if __name__ == "__main__":
    main()
